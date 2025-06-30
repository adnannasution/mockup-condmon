# import os
# import openpyxl
# import json

# # Buka Excel
# file_path = os.path.expanduser("~/Downloads/data_realtime.xlsm")
# wb = openpyxl.load_workbook(file_path, data_only=True)
# sheet = wb["GetHistorian"]
# value = sheet["F9"].value

# # Simpan ke file JSON (di folder yang sama dengan HTML kamu)
# with open("output.json", "w") as f:
#     json.dump({"vrms": value}, f)



import os
import openpyxl
import json
import time

# Path ke file Excel
file_path = os.path.expanduser("~/Downloads/data_realtime.xlsm")
output_file = "output.json"

while True:
    try:
        # Buka file Excel
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb["GetHistorian"]

        # Ambil tagnumber dari F5
        tagnumber = sheet["F5"].value

        # Cari nilai terakhir (terisi) di kolom F
        last_row = sheet.max_row
        vrms = None
        for row in range(last_row, 0, -1):
            value = sheet[f"F{row}"].value
            if value is not None:
                vrms = value
                break

        # Buat data JSON
        data = {
            "tagnumber": tagnumber,
            "vrms": vrms
        }

        # Simpan ke file output.json
        with open(output_file, "w") as f:
            json.dump(data, f)

        print(f"[UPDATE] {data}")
    except Exception as e:
        print(f"[ERROR] {e}")

    time.sleep(10)  # Update setiap 10 detik

